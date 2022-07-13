#Programmed by tg:@m0ma1a

import sqlite3
import openpyxl
import os
import telebot
from telebot import types
from decimal import *

from decouple import config

API_TOKEN = config('API_TOKEN')

base = sqlite3.connect("db.db")
cur = base.cursor()
cur.execute("""CREATE TABLE IF NOT EXiSTS list (
    curss TEXT
)""")
cur.execute("""CREATE TABLE IF NOT EXiSTS users (
    id TEXT
)""")
base.commit()
base.close()

bot = telebot.TeleBot(API_TOKEN)
@bot.message_handler(commands=['start'])
def syrrt(message):
    base = sqlite3.connect("db.db")
    cur = base.cursor()
    curs = cur.execute("SELECT curss FROM list").fetchone()[0]
    kb_m1 = types.InlineKeyboardMarkup(row_width=1)
    km1 = types.InlineKeyboardButton(text="Меню",callback_data="mn")
    km2 = types.InlineKeyboardButton(text = "Наша группа в VK",url="https://vk.cc/ceYh39")
    km3 = types.InlineKeyboardButton(text = "Наши отзывы в VK",url = "http://vk.cc/9IZVTn")
    kb_m1.add(km1,km3,km2)
    try:
        freq = cur.execute(f"SELECT * FROM users WHERE id='{message.chat.id}'").fetchone()[0]
    except:
        cur.execute(f"INSERT INTO users VALUES({message.chat.id})")
        base.commit()
        base.close
    #Programmed by tg:@m0ma1a
    bot.send_message(message.chat.id, f"Приветствую, {message.from_user.username} 🙋\nАктуальный курс - {curs}!", reply_markup=kb_m1)
def currrrs(message):
    cursa = message.text
    base = sqlite3.connect("db.db")
    cur = base.cursor()
    cur.execute(f"UPDATE list SET curss = {cursa}")
    base.commit()
    base.close()
    bot.send_message(message.chat.id,"Курс успешно измнен")
    base = sqlite3.connect("db.db")
    cur = base.cursor()
    matras =cur.execute("SELECT * FROM users").fetchall()
    for i in matras:
                #Programmed by tg:@m0ma1a
                bot.send_message(chat_id=i[0], text=f"Актуальный курс - {cursa}\nДля оформления заказа - vk.cc/cf1sGZ\nСпасибо, что остаетесь с нами.")
@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    if call.data == "mn":
        kb_m = types.ReplyKeyboardMarkup(row_width = 1,resize_keyboard=True)
        kb1 = types.KeyboardButton(text="Рассчитать стоимость заказа")
        kb2 = types.KeyboardButton(text="Узнать статус заказа")
        kb3 = types.KeyboardButton(text="Частые вопросы по заказам")
        kb_m.add(kb1,kb2,kb3)
        bot.send_message(call.message.chat.id, "Вы попали в Меню\nПожалуйста используйте кнопки ниже",reply_markup=kb_m)
    elif call.data == "curr":
        msg = bot.send_message(call.message.chat.id,"Отправьте Новый курс\nУчитывайте, что нужно использовать . а не ,")
        bot.register_next_step_handler(msg,currrrs)
    elif call.data == "ex":
        msg = bot.send_message(call.message.chat.id,"Отправьте exel таблицу")
        bot.register_next_step_handler(msg,dooccc)
       #Programmed by tg:@m0ma1a
def zakazst(message):
    try:
        zakaz = message.text
        base = sqlite3.connect("db.db")
        cur = base.cursor()
        stt = cur.execute(f"SELECT stat FROM cars WHERE brand ='{zakaz}'").fetchone()[0]
        kb_m = types.ReplyKeyboardMarkup(row_width = 1,resize_keyboard=True)
        kb1 = types.KeyboardButton(text="Рассчитать стоимость заказа")
        kb2 = types.KeyboardButton(text="Узнать статус заказа")
        kb3 = types.KeyboardButton(text="Частые вопросы по заказам")
        kb_m.add(kb1,kb2,kb3)
        bot.send_message(message.chat.id,f"Актуальный статус вашего заказа - {stt}",reply_markup=kb_m)
    except:
        bot.send_message(message.chat.id, "Произошла ошибка, пожалуйста проверьте номер вашего заказа")
@bot.message_handler(commands=['admin'])#Programmed by tg:@m0ma1a
def aaaaadd(message):
    if message.chat.id == "638426325":
        kb_ad = types.InlineKeyboardMarkup(row_width=1)
        l1 = types.InlineKeyboardButton(text="Добавить таблицу exel",callback_data="ex")
        l3 =  types.InlineKeyboardButton(text="Сменить Курс",callback_data="curr")
        kb_ad.add(l3,l1)
        bot.send_message(message.chat.id, "Меню админа\nПожалуйста используйте кнопки ниже",reply_markup=kb_ad)
    else:
        pass

def dooccc(message):
    connect = sqlite3.connect("db.db")
    cursor = connect.cursor()
    save_dir = os.getcwd()
    cursor.execute("DELETE FROM cars")
    connect.commit()#Programmed by tg:@m0ma1a
    file_name = message.document.file_name
    file_info = bot.get_file(message.document.file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    with open(file_name, 'wb') as new_file:
        new_file.write(downloaded_file)

    print('11111')
    prj_dir = os.path.abspath(os.path.curdir)

    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Имя базы


    # создание таблицы если ее не существует
    cursor.execute('CREATE TABLE IF NOT EXISTS cars (brand text, stat text)')
    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('Cars.xlsx', data_only=True)
    sheet = file_to_read['Sheet1']

    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []#Programmed by tg:@m0ma1a
        # Цикл по столбцам от 1 до 4 ( 5 не включая)
        for col in range(1, 3):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)

    # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        cursor.execute("INSERT INTO cars VALUES (?, ?);", (data[0], data[1]))
    connect.commit()
# закрытие соединения
    connect.close()#Programmed by tg:@m0ma1a
    base = sqlite3.connect("db.db")
    cur = base.cursor()
    matras =cur.execute("SELECT * FROM users").fetchall()
    for i in matras:
                print(i)
                bot.send_message(chat_id=i[0], text=f"Мы обновили статус для каждого заказа - проверяйте!")

@bot.message_handler(content_types=['text'])
def adaddad(message):
    if message.text == "Рассчитать стоимость заказа":
        a = types.ReplyKeyboardRemove()
        msg = bot.send_message(message.chat.id, "Укажите стоимость товара в юанях",reply_markup=a)
        bot.register_next_step_handler(msg,chenik)
    elif message.text == "Узнать статус заказа":#Programmed by tg:@m0ma1a
        msg = bot.send_message(message.chat.id, "Укажите номер заказа")
        bot.register_next_step_handler(msg,zakazst)
    elif message.text == "Группа ВК":
        kb_vk = types.InlineKeyboardMarkup()
        k1 = types.InlineKeyboardButton(text="Группа ВК",url="https://vk.cc/ceYh39")
        kb_vk.add(k1)
        bot.send_message(message.chat.id,"Подпищись на нашу группу в вконтакте там выходят новости каждый день",reply_markup=kb_vk)
    elif message.text == "Назад":
        kb_m = types.ReplyKeyboardMarkup(row_width = 1,resize_keyboard=True)
        kb1 = types.KeyboardButton(text="Рассчитать стоимость заказа")
        kb2 = types.KeyboardButton(text="Узнать статус заказа")
        kb3 = types.KeyboardButton(text="Частые вопросы по заказам")
        kb_m.add(kb1,kb2,kb3)#Programmed by tg:@m0ma1a
        bot.send_message(message.chat.id, "Вы попали в Меню\nПожалуйста используйте кнопки ниже",reply_markup=kb_m)
    elif message.text == "Частые вопросы по заказам":
        kb_a = types.ReplyKeyboardMarkup(row_width=2,resize_keyboard=True)
        kb1a = types.KeyboardButton(text="Оригинальная продукция?")
        kb2a = types.KeyboardButton(text="Какой срок доставки?")
        kb3a = types.KeyboardButton(text="Где можно забрать свой заказ?")
        kb4a = types.KeyboardButton(text="Могу ли я забрать заказ курьером?")
        kb5a = types.KeyboardButton(text="Можно ли доставить заказ в другой город?")
        p1 = types.KeyboardButton(text="Назад")
        kb_a.row(kb1a).add(kb2a,kb3a).add(kb4a,kb5a).row(p1)
        bot.send_message(message.chat.id,"Выберете пункт в меню ниже",reply_markup=kb_a)
    elif message.text == "Отзывы" :#Programmed by tg:@m0ma1a
        kb_ot = types.InlineKeyboardMarkup()
        k11 = types.InlineKeyboardButton(text="Отзывы",url="http://vk.cc/9IZVTn")
        kb_ot.add(k11)#Programmed by tg:@m0ma1a
        bot.send_message(message.chat.id,"Отзывы о нашем проекте вы можете посмотреть тут",reply_markup=kb_ot)
    elif message.text == "Оригинальная продукция?":
        bot.send_message(message.chat.id,f"Да, исключительно оригинальные и аутентичные товары, которые проходят проверку на складе Poizon.\nЛюбые проверки с вашей стороны!")
    elif message.text == "Какой срок доставки?":
        bot.send_message(message.chat.id,f"Ориентировочный срок доставки 14 - 20 дней не включая доставку с приложения.\nСрок доставки варьируется от выбранной кнопки в приложении.\nУказанные сроки являются приблизительными, и могут отличаться как в меньшую, так и в большую сторону.")
    elif message.text == "Где можно забрать свой заказ?":
        bot.send_message(message.chat.id,f"После подтверждения,что ваш заказ готов к выдаче, его можно забрать по следующему адресу:\nм.Южная\nг.Москва,ул.Сумская,д.12/17,п.7")
    elif message.text == "Могу ли я забрать заказ курьером?":
        bot.send_message(message.chat.id,f"Да, вы можете забрать свой заказ курьером предварительно связавшись с нами!")
    elif message.text == "Можно ли доставить заказ в другой город?":
        bot.send_message(message.chat.id,f"Да, можно! Мы можем отправить следующими курьерскими службами на ваш выбор:\nСДЭК или Почта России.\nДля этого укажите свои данные для отправки при оформлении заказа.")

#Programmed by tg:@m0ma1a

def chenik(message):
    base = sqlite3.connect("db.db")
    cur = base.cursor()
    curs = cur.execute("SELECT curss FROM list").fetchone()[0]
    try:

        if int(message.text) <= 2000:
            si = Decimal(int(message.text)+28)*Decimal(curs)//1
            proc = si / 100 * 5
            if proc <= 500:
                proc = 500
            else:
                pass

            kb_m = types.ReplyKeyboardMarkup(row_width = 1,resize_keyboard=True)
            kb1 = types.KeyboardButton(text="Рассчитать стоимость заказа")
            kb2 = types.KeyboardButton(text="Узнать статус заказа")
            kb3 = types.KeyboardButton(text="Частые вопросы по заказам")
            kb_m.add(kb1,kb2,kb3)
            bot.send_message(message.chat.id,f"Актуальный Курс - {curs}\nИтоговая стоимость заказа - {si+proc//1}\nЦена указана без учета стоимости доставки.\nДоставка оплачивается при получении товара.\nСтоимость доставки - 130 юаней / кг!",reply_markup=kb_m)
        elif int(message.text) > 2000 and int(message.text)<4000:
            si = Decimal(int(message.text)+28)*Decimal(curs)//1
            proc = si / 100 * 4
            if proc < 1000:
                proc = 1000
            else:
                pass#Programmed by tg:@m0ma1a
            kb_m = types.ReplyKeyboardMarkup(row_width = 1,resize_keyboard=True)
            kb1 = types.KeyboardButton(text="Рассчитать стоимость заказа")
            kb2 = types.KeyboardButton(text="Узнать статус заказа")
            kb3 = types.KeyboardButton(text="Частые вопросы по заказам")
            kb_m.add(kb1,kb2,kb3)
            bot.send_message(message.chat.id,f"Актуальный Курс - {curs}\nИтоговая стоимость заказа - {si+proc//1}\nЦена указана без учета стоимости доставки.\nДоставка оплачивается при получении товара.\nСтоимость доставки - 130 юаней / кг!",reply_markup=kb_m)
        elif int(message.text) > 4000:
            si = Decimal(int(message.text)+28)*Decimal(curs)//1
            proc = si / 1000 * 35
            print(proc)
            proc = proc /100
            if proc < 1600:
                proc = 1600
            else:
                pass#Programmed by tg:@m0ma1a
            kb_m = types.ReplyKeyboardMarkup(row_width = 1,resize_keyboard=True)
            kb1 = types.KeyboardButton(text="Рассчитать стоимость заказа")
            kb2 = types.KeyboardButton(text="Узнать статус заказа")
            kb3 = types.KeyboardButton(text="Частые вопросы по заказам")
            kb_m.add(kb1,kb2,kb3)
            bot.send_message(message.chat.id,f"Актуальный Курс - {curs}\nИтоговая стоимость заказа - {si+proc//1}\nЦена указана без учета стоимости доставки.\nДоставка оплачивается при получении товара.\nСтоимость доставки - 130 юаней / кг!",reply_markup=kb_m)
    except:#Programmed by tg:@m0ma1a
        bot.send_message(message.chat.id, "Неверная команда")
    base.close()
bot.polling(non_stop=True)#Programmed by tg:@m0ma1a

#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a
#Programmed by tg:@m0ma1a